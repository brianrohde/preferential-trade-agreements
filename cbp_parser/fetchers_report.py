import os
import pandas as pd
from typing import List, Dict
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from .tier_fetchers import fetch_tier_1, fetch_tier_2, fetch_tier_3
from .utils import ensure_dir


def has_line_breaks(text: str) -> bool:
    return "\n" in text or "\r" in text


def run_all_tiers(ruling_ids: List[str], cache_dir: str) -> List[Dict]:
    results = []
    
    for ruling_id in ruling_ids:
        print(f"\nFetching all tiers for {ruling_id}...")
        
        for tier_num, tier_name, fetch_func in [
            (1, "tier_1 - JSON API", fetch_tier_1),
            (2, "tier_2 - HTML Page", fetch_tier_2),
            (3, "tier_3 - Document Download", fetch_tier_3),
        ]:
            try:
                text, pretty, meta = fetch_func(ruling_id, cache_dir)
                results.append({
                    "ruling_id": ruling_id,
                    "tier": tier_name,
                    "text_length": len(text),
                    "has_line_breaks": has_line_breaks(text),
                    "status": "Success"
                })
                print(f"  ✓ {tier_name}: {len(text)} chars, line_breaks={has_line_breaks(text)}")
            except Exception as e:
                results.append({
                    "ruling_id": ruling_id,
                    "tier": tier_name,
                    "text_length": 0,
                    "has_line_breaks": False,
                    "status": f"Failed: {str(e)[:50]}"
                })
                print(f"  ✗ {tier_name} failed: {e}")
    
    return results


def export_fetchers_report(results: List[Dict], output_path: str) -> None:
    df = pd.DataFrame(results)
    df = df[["ruling_id", "tier", "text_length", "has_line_breaks", "status"]]
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="tier_comparison", index=False)
        ws = writer.sheets["tier_comparison"]
        
        header_font = Font(color="FFFFFF", bold=True)
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
        
        lb_col = df.columns.get_loc("has_line_breaks") + 1
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=lb_col)
            if cell.value == True:
                cell.fill = green_fill
            elif cell.value == False:
                cell.fill = red_fill
        
        tl_col = df.columns.get_loc("text_length") + 1
        tl_col_letter = get_column_letter(tl_col)
        tl_range = f"{tl_col_letter}2:{tl_col_letter}{len(df) + 1}"
        
        ws.conditional_formatting.add(
            tl_range,
            ColorScaleRule(
                start_type="min", start_color="FFC7CE",
                mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                end_type="max", end_color="C6EFCE"
            )
        )
        
        for idx, col in enumerate(df.columns, 1):
            max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
            max_length = min(max_length, 50)
            ws.column_dimensions[get_column_letter(idx)].width = max_length
    
    print(f"\n✓ Fetchers report exported: {output_path}")
