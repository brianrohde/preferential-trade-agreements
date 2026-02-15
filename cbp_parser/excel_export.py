"""
CBP Parser - Excel Export Module v.2
=================================

Exports triage results to Excel workbook with three sheets:

1. **summary** — One row per ruling_id with disagreement metrics
2. **details** — Long format: ruling_id | extraction_type | url | fields
3. **metadata** — Timestamp and script execution settings

Features:
- Excel table formatting with filters and color schemes
- Conditional formatting (Yes=Green, No=Red, disagreement counts)
- Proper HTML formatting handling (replying_person with line breaks)
- Informational placeholders when benchmark missing (e.g., "N/A", "No Bench")
- Correct column ordering (ruling_id first, then extraction_type, then url)

Usage (from extract_rulings.py):
    from cbp_parser.excel_export import export_to_excel
    export_to_excel(
        output_path="out/04_review/review.xlsx",
        triage=triage_dict,
        bench_spec=bench_spec,
        regex_records=regex_results,
        llm_records=llm_results,
        bench_values=benchmark_data,
        llm_enabled=True,  # NEW: whether --llm flag was used
        timestamp=datetime.now()  # NEW: when extraction ran
    )
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime
from openpyxl.styles import Alignment


def _normalize_bench_values(bench_values: Optional[Dict | List]) -> Dict:
    """
    Convert benchmark values to dict keyed by ruling_id.
    
    Handles both list and dict formats from load_benchmark_values().
    
    Args:
        bench_values: Either list of records or dict {ruling_id: record}
        
    Returns:
        Dict keyed by ruling_id
    """
    if bench_values is None:
        return {}
    
    if isinstance(bench_values, list):
        return {r.get("ruling_id"): r for r in bench_values if isinstance(r, dict) and r.get("ruling_id")}
    elif isinstance(bench_values, dict):
        return bench_values
    else:
        return {}


def _get_ruling_url(ruling_id: str) -> str:
    """
    Generate CBP ruling lookup URL for manual inspection.
    
    Args:
        ruling_id: Internal ruling ID
        
    Returns:
        URL string to CBP Customs Rulings Issuances & Decisions database
    """
    base_url = "https://rulings.cbp.gov/search?term="
    return f"{base_url}{ruling_id}"


def _convert_br_to_newlines(text: Optional[str]) -> str:
    """
    Convert HTML <br> tags to Excel-friendly line breaks.

    Requirement:
    - Insert a blank line between each original line -> double line break.
    - Keep wrap OFF and default row heights (visual display may show only first line).
    """
    if text is None or not isinstance(text, str):
        return text

    # Normalize all <br> variants to '\n'
    t = text.replace("<br />", "\n").replace("<br/>", "\n").replace("<br>", "\n")

    # Normalize existing CRLF/LF to LF, collapse any runs, then enforce blank lines
    # Split on any line breaks, strip each line, drop empties
    lines = [ln.strip() for ln in t.replace("\r\n", "\n").split("\n") if ln.strip()]

    # Use CRLF because Excel sometimes only renders properly after edit with LF-only
    # And use DOUBLE CRLF to enforce a blank line between entries
    return "\r\n\r\n".join(lines)



def _style_header_row(ws, max_col: int) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="top", wrap_text=False)

    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align


def _build_summary_df(triage: Dict, bench_values_dict: Dict) -> pd.DataFrame:
    """
    Build summary sheet: one row per ruling_id with disagreement counts.
    
    Args:
        triage: Triage report dict from triage_report_goal()
        bench_values_dict: Normalized benchmark dict {ruling_id: record}
        
    Returns:
        DataFrame with columns: ruling_id, has_bench, has_llm, 
                               n_disagree_regex_llm, disagree_fields_regex_llm, etc.
    """
    rows = []
    
    # Pre-compute set of ruling IDs with benchmark
    bench_ruling_ids = set(bench_values_dict.keys())
    
    for ruling_id, field_dict in triage.items():
        row = {"ruling_id": ruling_id}
        
        # Check if benchmark exists for this ruling
        has_bench = ruling_id in bench_ruling_ids
        row["has_bench"] = "Yes" if has_bench else "No"
        
        # Track which extraction types are present
        has_llm = any(
            field_dict.get(f, {}).get("llm") is not None
            for f in field_dict
        )
        row["has_llm"] = "Yes" if has_llm else "No"
        
        # Count disagreements between extraction types
        regex_llm_disagree = []
        regex_bench_disagree = []
        llm_bench_disagree = []
        
        for field, values in field_dict.items():
            regex_val = values.get("regex")
            llm_val = values.get("llm")
            bench_val = values.get("bench")
            
            # regex vs llm comparison
            if regex_val != llm_val and llm_val is not None:
                regex_llm_disagree.append(field)
            
            # regex vs bench comparison (only if bench exists)
            if has_bench and regex_val != bench_val and bench_val is not None:
                regex_bench_disagree.append(field)
            
            # llm vs bench comparison (only if bench exists and llm exists)
            if has_bench and llm_val is not None and llm_val != bench_val and bench_val is not None:
                llm_bench_disagree.append(field)
        
        row["n_disagree_regex_llm"] = len(regex_llm_disagree)
        row["disagree_fields_regex_llm"] = "; ".join(regex_llm_disagree) if regex_llm_disagree else "N/A"
        
        # If no bench, show "No Bench" instead of 0
        if has_bench:
            row["n_disagree_regex_bench"] = len(regex_bench_disagree)
            row["disagree_fields_regex_bench"] = "; ".join(regex_bench_disagree) if regex_bench_disagree else "N/A"
            row["n_disagree_llm_bench"] = len(llm_bench_disagree)
            row["disagree_fields_llm_bench"] = "; ".join(llm_bench_disagree) if llm_bench_disagree else "N/A"
        else:
            row["n_disagree_regex_bench"] = "No Bench"
            row["disagree_fields_regex_bench"] = "No Bench"
            row["n_disagree_llm_bench"] = "No Bench"
            row["disagree_fields_llm_bench"] = "No Bench"
        
        rows.append(row)
    
    df_summary = pd.DataFrame(rows)
    df_summary = df_summary.sort_values("ruling_id").reset_index(drop=True)
    
    return df_summary


def _build_details_df(
    triage: Dict,
    bench_spec: Dict,
    regex_records: List[Dict],
    llm_records: List[Dict],
    bench_values_dict: Dict
) -> pd.DataFrame:
    """
    Build details sheet in long format: one row per (extraction_type, ruling_id).
    
    Column order: ruling_id | extraction_type | url | field_1 | ... | field_N
    
    Args:
        triage: Triage dict
        bench_spec: Benchmark specification with field_order
        regex_records: List of regex extraction results
        llm_records: List of LLM extraction results
        bench_values_dict: Normalized benchmark dict {ruling_id: record}
        
    Returns:
        DataFrame in long format with proper column ordering
    """
    field_order = bench_spec["output"]["field_order"]

    # Avoid duplicate columns: we already add ruling_id/extraction_type/url explicitly
    reserved = {"ruling_id", "extraction_type", "url"}
    field_order = [f for f in field_order if f not in reserved]

    # Build lookup dicts for faster access
    regex_by_id = {r.get("ruling_id"): r for r in regex_records}
    llm_by_id = {r.get("ruling_id"): r for r in llm_records}
    
    rows = []
    
    # Iterate each ruling in triage
    for ruling_id in sorted(triage.keys()):
        url = _get_ruling_url(ruling_id)
        
        # Row 1: Regex extraction
        regex_record = regex_by_id.get(ruling_id, {})
        regex_row = {
            "ruling_id": ruling_id,
            "extraction_type": "regex",
            "url": url
        }
        for field in field_order:
            val = regex_record.get(field)
            # Convert <br> to actual newlines for better Excel display
            if field == "replying_person" and val:
                val = _convert_br_to_newlines(val)
            regex_row[field] = val if val is not None else "[No Data Extracted]"
        rows.append(regex_row)
        
        # Row 2: LLM extraction (if available)
        llm_record = llm_by_id.get(ruling_id)
        if llm_record is not None:
            llm_row = {
                "ruling_id": ruling_id,
                "extraction_type": "llm",
                "url": url
            }
            for field in field_order:
                val = llm_record.get(field)
                if field == "replying_person" and val:
                    val = _convert_br_to_newlines(val)
                llm_row[field] = val if val is not None else "[No Data Extracted]"
            rows.append(llm_row)
        
        # Row 3: Benchmark ground truth (if available)
        bench_record = bench_values_dict.get(ruling_id)
        if bench_record is not None:
            bench_row = {
                "ruling_id": ruling_id,
                "extraction_type": "bench",
                "url": url
            }
            for field in field_order:
                val = bench_record.get(field)
                if field == "replying_person" and val:
                    val = _convert_br_to_newlines(val)
                bench_row[field] = val if val is not None else "[Benchmark Missing]"
            rows.append(bench_row)
    
    df_details = pd.DataFrame(rows)


    df_details = df_details.copy()
    
    df_details.columns = _dedupe_columns(df_details.columns)

    # Ensure column order: ruling_id, extraction_type, url, then all fields
    col_order = ["ruling_id", "extraction_type", "url"] + field_order
    df_details = df_details[col_order]
    
    return df_details

def _build_data_dictionary_df(bench_spec: Dict) -> pd.DataFrame:
    """
    Human-facing dictionary of fields across sheets + important notes.
    """
    field_order = bench_spec["output"]["field_order"]

    rows = []

    # Summary sheet fields (fixed)
    summary_fields = [
        ("summary", "ruling_id", "Internal ruling identifier", ""),
        ("summary", "has_bench", "Whether benchmark record exists for ruling_id", "Yes/No (conditional formatting)."),
        ("summary", "has_llm", "Whether LLM extraction is available for ruling_id", "Yes/No (conditional formatting)."),
        ("summary", "n_disagree_regex_llm", "Count of fields where regex != llm", "Color scale conditional formatting."),
        ("summary", "disagree_fields_regex_llm", "Semicolon-separated fields where regex != llm", ""),
        ("summary", "n_disagree_regex_bench", "Count of fields where regex != bench", "\"No Bench\" when benchmark missing."),
        ("summary", "disagree_fields_regex_bench", "Semicolon-separated fields where regex != bench", "\"No Bench\" when benchmark missing."),
        ("summary", "n_disagree_llm_bench", "Count of fields where llm != bench", "\"No Bench\" when benchmark missing."),
        ("summary", "disagree_fields_llm_bench", "Semicolon-separated fields where llm != bench", "\"No Bench\" when benchmark missing."),
    ]
    for r in summary_fields:
        rows.append(r)

    # Details sheet fields
    rows.append(("details", "ruling_id", "Internal ruling identifier", "Use filter by ruling_id to compare methods."))
    rows.append(("details", "extraction_type", "Which method produced the row", "Values: regex | llm | bench."))
    rows.append(("details", "url", "CBP search URL for manual review", ""))

    for f in field_order:
        note = ""
        if f == "replying_person":
            note = (
                "Value contains multiple sub-lines separated by a DOUBLE line break "
                "(Char(10)&Char(10)). With Wrap Text off, Excel may not display all lines, "
                "but the underlying value is correct for splitting."
            )
        rows.append(("details", f, "Extracted field", note))

    # Metadata sheet fields
    metadata_fields = [
        ("metadata", "Execution Timestamp", "When the extraction run executed", ""),
        ("metadata", "LLM Extraction Enabled", "Whether --llm flag was used this run", "Yes/No (conditional formatting)."),
        ("metadata", "LLM Extraction Available", "Whether LLM rows exist (possibly from prior run)", "Yes/No (conditional formatting)."),
        ("metadata", "LLM Extraction Updated", "Whether new LLM results were generated this run", "Yes/No (conditional formatting)."),
        ("metadata", "Total Rulings Processed", "Count of rulings processed (regex)", "Stored as number."),
        ("metadata", "Rulings with LLM Results", "Count of rulings that have LLM output", "Stored as number."),
        ("metadata", "Rulings without LLM Results", "Count of rulings missing LLM output", "Stored as number."),
    ]
    for r in metadata_fields:
        rows.append(r)

    return pd.DataFrame(rows, columns=["sheet", "field", "meaning", "notes"])


def _build_metadata_df(
    regex_records: List[Dict],
    llm_records: List[Dict],
    llm_enabled: bool,
    llm_updated_this_run: bool,
    timestamp: Optional[datetime] = None
) -> pd.DataFrame:
    """
    Build metadata sheet with execution info and summary stats.
    
    Args:
        regex_records: List of regex extraction results
        llm_records: List of LLM extraction results
        llm_enabled: Whether --llm flag was used
        timestamp: When extraction was run
        
    Returns:
        DataFrame with metadata (key-value pairs)
    """
    if timestamp is None:
            timestamp = datetime.now()

    llm_available = len(llm_records) > 0

    metadata = {
        "Key": [
            "Execution Timestamp",
            "LLM Extraction Enabled",
            "LLM Extraction Available",
            "LLM Extraction Updated",
            "Total Rulings Processed",
            "Rulings with LLM Results",
            "Rulings without LLM Results",
        ],
        "Value": [
            timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            "Yes" if llm_enabled else "No",
            "Yes" if llm_available else "No",
            "Yes" if llm_updated_this_run else "No",
            len(regex_records),
            len(llm_records),
            len(regex_records) - len(llm_records),
        ],
    }

    return pd.DataFrame(metadata)

def _dedupe_columns(cols):
    seen = {}
    out = []
    for c in cols:
        if c not in seen:
            seen[c] = 0
            out.append(c)
        else:
            seen[c] += 1
            out.append(f"{c}_{seen[c]}")
    return out


def _apply_table_formatting(writer, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Apply Excel table formatting to a sheet.
    
    Converts data range to a named Excel table with:
    - Filter buttons on header row
    - Predefined color scheme
    - Auto-sized columns
    
    Args:
        writer: pd.ExcelWriter object
        sheet_name: Name of the sheet
        df: DataFrame that was written to the sheet
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    worksheet = writer.sheets[sheet_name]
    
    # Define table range: all data including header
    max_row = len(df) + 1  # +1 for header
    max_col = len(df.columns)
    _style_header_row(worksheet, max_col)

    # Build range string (e.g., "A1:Z100")
    from openpyxl.utils import get_column_letter
    start_col = get_column_letter(1)
    end_col = get_column_letter(max_col)
    table_range = f"{start_col}1:{end_col}{max_row}"
    
    # Create table
    tab = Table(displayName=f"Table_{sheet_name}", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleLight8",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True, 
        showColumnStripes=True    
    )
    tab.tableStyleInfo = style
    tab.totalsRowShown = True

    worksheet.add_table(tab)
    
    # Auto-size columns
    for idx, col in enumerate(df.columns, 1):
        max_length = max(
            df[col].astype(str).map(len).max(),
            len(col)
        ) + 2
        
        # Cap at 50 for readability (some fields are very long)
        max_length = min(max_length, 50)
        
        col_letter = get_column_letter(idx)
        worksheet.column_dimensions[col_letter].width = max_length
        
        # Adjust the text alignment
        top_left_no_wrap = Alignment(horizontal="left", vertical="top", wrap_text=False)

        for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = top_left_no_wrap


def _apply_conditional_formatting(writer) -> None:
    """
    Apply Excel-native conditional formatting rules (dynamic, updates after edits).

    - summary: has_bench/has_llm Yes=green, No=red (as CF rules, not static fills)
    - summary: n_disagree_* columns get 3-color scale
    - metadata: Yes/No values in Value column get Yes=green, No=red (CF rules)
    """
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
    from openpyxl.utils import get_column_letter

    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Excel-ish light green
    no_fill  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Excel-ish light red

    # -------------------------
    # Summary sheet formatting
    # -------------------------
    if "summary" in writer.sheets:
        ws = writer.sheets["summary"]
        header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        def _col_index(col_name: str) -> int | None:
            try:
                return header.index(col_name) + 1
            except ValueError:
                return None

        # Yes/No CF rules for has_bench and has_llm
        for col_name in ("has_bench", "has_llm"):
            col_idx = _col_index(col_name)
            if not col_idx:
                continue
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"

            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="equal", formula=['"Yes"'], fill=yes_fill)
            )
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="equal", formula=['"No"'], fill=no_fill)
            )

        # Color scales for all numeric disagreement columns
        # IMPORTANT: do NOT use the old buggy col_name[11].isdigit() logic
        for col_idx, col_name in enumerate(header, 1):
            if not col_name:
                continue
            if str(col_name).startswith("n_disagree_"):
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"

                # Only apply if cells are numeric; your "No Bench" strings will simply not participate
                ws.conditional_formatting.add(
                    cell_range,
                    ColorScaleRule(
                        start_type="min", start_color="C6EFCE",   # green low
                        mid_type="percentile", mid_value=50, mid_color="FFEB9C",  # yellow mid
                        end_type="max", end_color="FFC7CE"       # red high
                    )
                )

    # -------------------------
    # Metadata sheet formatting
    # -------------------------
    if "metadata" in writer.sheets:
        ws = writer.sheets["metadata"]

        # We expect 2 columns: Key | Value
        # Apply Yes/No rules to entire Value column (col B), rows 2..max
        cell_range = f"B2:B{ws.max_row}"

        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"Yes"'], fill=yes_fill)
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"No"'], fill=no_fill)
        )

def _apply_static_detail_diff_highlighting(writer, df_details: pd.DataFrame) -> None:
    """
    Static highlighting rules (NOT conditional formatting):
    - If bench exists for ruling and field differs vs bench: highlight bench cell light green.
    - Else if no bench and regex vs llm differ: highlight regex light red, llm light yellow.
    - No highlight if no disagreement.
    """
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    ws = writer.sheets["details"]

    fill_bench = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
    fill_regex = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red
    fill_llm   = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # light yellow

    # Map column name -> excel column index
    cols = list(df_details.columns)
    col_index = {c: i + 1 for i, c in enumerate(cols)}

    key_cols = {"ruling_id", "extraction_type", "url"}
    value_cols = [c for c in cols if c not in key_cols]

    # Build per-ruling row index lookup: ruling_id -> {extraction_type: excel_row_number}
    # df_details row 0 is Excel row 2 (row 1 is header)
    by_ruling = {}
    for i, row in df_details.reset_index(drop=True).iterrows():
        rid = row["ruling_id"]
        et = row["extraction_type"]
        excel_row = i + 2
        by_ruling.setdefault(rid, {})[et] = excel_row

    # For each ruling, decide which comparisons to apply
    for rid, row_map in by_ruling.items():
        bench_row = row_map.get("bench")
        regex_row = row_map.get("regex")
        llm_row   = row_map.get("llm")

        has_bench = bench_row is not None
        has_llm = llm_row is not None

        for field in value_cols:
            c = col_index[field]

            # Pull values (None if row missing)
            bench_val = ws.cell(row=bench_row, column=c).value if has_bench else None
            regex_val = ws.cell(row=regex_row, column=c).value if regex_row else None
            llm_val   = ws.cell(row=llm_row, column=c).value if llm_row else None

            # ---- Case 1: bench exists ----
            if has_bench:
                bench_cell = ws.cell(row=bench_row, column=c)
                bench_cell.fill = fill_bench  # bench is always "as it should be"

                # If regex exists: green if matches bench else red
                if regex_row:
                    regex_cell = ws.cell(row=regex_row, column=c)
                    if regex_val == bench_val:
                        regex_cell.fill = fill_bench
                    else:
                        regex_cell.fill = fill_regex

                # If llm exists: green if matches bench else yellow
                if llm_row:
                    llm_cell = ws.cell(row=llm_row, column=c)
                    if llm_val == bench_val:
                        llm_cell.fill = fill_bench
                    else:
                        llm_cell.fill = fill_llm

            # ---- Case 2: no bench ----
            else:
                # Only color if both exist
                if regex_row and llm_row:
                    regex_cell = ws.cell(row=regex_row, column=c)
                    llm_cell = ws.cell(row=llm_row, column=c)

                    if regex_val == llm_val:
                        regex_cell.fill = fill_bench  # green
                        llm_cell.fill = fill_bench    # green
                    else:
                        regex_cell.fill = fill_regex  # red
                        llm_cell.fill = fill_llm      # yellow
                        
    # Color-code extraction_type column (static fills)
    etype_col = col_index.get("extraction_type")
    if etype_col:
        for i, row in df_details.reset_index(drop=True).iterrows():
            excel_row = i + 2  # header is row 1
            et = row["extraction_type"]

            cell = ws.cell(row=excel_row, column=etype_col)

            if et == "bench":
                cell.fill = fill_bench   # green
            elif et == "regex":
                cell.fill = fill_regex   # red
            elif et == "llm":
                cell.fill = fill_llm     # yellow/orange


def _highlight_data_dictionary_replying_person_note(writer) -> None:
    from openpyxl.styles import PatternFill, Font

    if "data_dictionary" not in writer.sheets:
        return

    ws = writer.sheets["data_dictionary"]

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    black_font = Font(color="000000")

    # Find the "field" and "notes" columns by header name
    header = [c.value for c in ws[1]]
    try:
        field_col = header.index("field") + 1
        notes_col = header.index("notes") + 1
    except ValueError:
        return

    # For each row where field == replying_person, highlight notes cell
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=field_col).value == "replying_person":
            cell = ws.cell(row=r, column=notes_col)
            cell.fill = yellow
            cell.font = black_font

def _autosize_notes_column(writer, sheet_name: str = "data_dictionary", max_width: int = 90) -> None:
    from openpyxl.utils import get_column_letter

    if sheet_name not in writer.sheets:
        return

    ws = writer.sheets[sheet_name]
    header = [c.value for c in ws[1]]
    try:
        notes_col = header.index("notes") + 1
    except ValueError:
        return

    # Compute max visible length in notes column
    max_len = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=notes_col).value
        if v is None:
            continue
        # Consider longest line if there are line breaks
        s = str(v)
        s = max(s.splitlines(), key=len) if "\n" in s or "\r" in s else s
        max_len = max(max_len, len(s))

    # Rough Excel width heuristic
    width = min(max(12, int(max_len * 1.1) + 2), max_width)
    ws.column_dimensions[get_column_letter(notes_col)].width = width

def _disable_gridlines(writer) -> None:
    for ws in writer.book.worksheets:
        ws.sheet_view.showGridLines = False


def export_to_excel(
    output_path: str,
    triage: Dict,
    bench_spec: Dict,
    regex_records: List[Dict],
    llm_records: Optional[List[Dict]] = None,
    bench_values: Optional[Dict | List] = None,
    llm_enabled: bool = False,
    llm_updated_this_run: bool = False,
    timestamp: Optional[datetime] = None
) -> None:
    """
    Export triage results to Excel workbook with summary, details, and metadata sheets.
    
    Args:
        output_path: Path to output .xlsx file (e.g., "out/04_review/review.xlsx")
        triage: Triage report dict from triage_report_goal()
        bench_spec: Benchmark specification with field definitions
        regex_records: Regex extraction results
        llm_records: LLM extraction results (optional, may be empty list)
        bench_values: Benchmark ground truth values (optional, can be list or dict)
        llm_enabled: Whether --llm flag was used (for metadata)
        timestamp: When extraction was run (defaults to now)
        
    Raises:
        ValueError: If output path is not .xlsx
        OSError: If parent directory cannot be created
    """
    # Validate output path
    output_path = Path(output_path)
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Output path must be .xlsx file, got: {output_path}")
    
    # Create parent directory if needed
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Default to empty collections if not provided
    if llm_records is None:
        llm_records = []
    
    # Normalize bench_values to dict
    bench_values_dict = _normalize_bench_values(bench_values)
    
    # Build DataFrames
    df_metadata = _build_metadata_df(
        regex_records=regex_records,
        llm_records=llm_records,
        llm_enabled=llm_enabled,
        llm_updated_this_run=llm_updated_this_run,
        timestamp=timestamp
    )
    df_dict = _build_data_dictionary_df(bench_spec)
    df_summary = _build_summary_df(triage, bench_values_dict)
    df_details = _build_details_df(
        triage=triage,
        bench_spec=bench_spec,
        regex_records=regex_records,
        llm_records=llm_records,
        bench_values_dict=bench_values_dict
    )

    

    
    # Write to Excel with three sheets
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_metadata.to_excel(writer, sheet_name="metadata", index=False)
        df_dict.to_excel(writer, sheet_name="data_dictionary", index=False)
        df_summary.to_excel(writer, sheet_name="summary", index=False)
        df_details.to_excel(writer, sheet_name="details", index=False)
        
        # Apply table formatting and styling
        _apply_table_formatting(writer, "metadata", df_metadata)
        _apply_table_formatting(writer, "data_dictionary", df_dict)
        _apply_table_formatting(writer, "summary", df_summary)

        print(df_details.columns[df_details.columns.duplicated()].tolist())
        _apply_table_formatting(writer, "details", df_details)
       
        # Apply conditional formatting
        _apply_conditional_formatting(writer)
        _apply_static_detail_diff_highlighting(writer, df_details)
        _highlight_data_dictionary_replying_person_note(writer)

        # Apply Other Stylistic Choices
        _disable_gridlines(writer)
        _autosize_notes_column(writer, "data_dictionary")

    
    print(f"✓ Excel export: {output_path}")
    print(f"  - summary: {len(df_summary)} ruling(s)")
    print(f"  - details: {len(df_details)} extraction record(s)")
    print(f"  - metadata: execution info")